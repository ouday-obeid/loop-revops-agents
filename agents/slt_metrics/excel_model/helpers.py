"""Excel layout + chart helpers used across every sheet."""
from __future__ import annotations

from typing import Any, Sequence

from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from agents.slt_metrics.excel_model import styles as S


# ---------------------------------------------------------------- layout

def freeze_header(ws: Worksheet, *, rows: int = 1) -> None:
    """Freeze top N rows (plus the first column) for easy scrolling."""
    ws.freeze_panes = ws.cell(row=rows + 1, column=1).coordinate


def auto_width(ws: Worksheet, *, min_w: int = 10, max_w: int = 60) -> None:
    """Set each column's width to the longest value + padding, clamped."""
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        longest = 0
        for cell in col:
            v = cell.value
            if v is None:
                continue
            longest = max(longest, len(str(v)))
        ws.column_dimensions[letter].width = max(min_w, min(max_w, longest + 2))


def write_header_row(ws: Worksheet, row: int, headers: Sequence[str]) -> None:
    for col_idx, name in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=name)
        cell.fill = S.FILL_HEADER
        cell.font = S.FONT_HEADER
        cell.alignment = S.ALIGN_HEADER
        cell.border = S.BORDER_CELL
    ws.row_dimensions[row].height = 24


def write_title_banner(ws: Worksheet, title: str, *, cols: int) -> None:
    """Merge row 1 across `cols` columns and stamp the workbook title."""
    ws.cell(row=1, column=1, value=title)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
    cell = ws.cell(row=1, column=1)
    cell.fill = S.FILL_TITLE
    cell.font = S.FONT_TITLE
    cell.alignment = S.ALIGN_CENTER
    ws.row_dimensions[1].height = 28


def write_body_row(
    ws: Worksheet,
    row: int,
    values: Sequence[Any],
    *,
    number_formats: Sequence[str | None] | None = None,
    zebra: bool = True,
) -> None:
    """Write a body row with zebra striping and optional per-column formats."""
    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        cell.font = S.FONT_BODY
        cell.border = S.BORDER_CELL
        if zebra and row % 2 == 0:
            cell.fill = S.FILL_ALT_ROW
        if number_formats and col_idx - 1 < len(number_formats):
            fmt = number_formats[col_idx - 1]
            if fmt:
                cell.number_format = fmt


def write_gap_row(ws: Worksheet, row: int, cols: int) -> None:
    """Emit a single row of '-- (Loop Pulse unavailable)' across `cols` cells."""
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c, value=S.GAP_TEXT)
        cell.font = S.FONT_GAP
        cell.fill = S.FILL_GAP
        cell.border = S.BORDER_CELL


# ---------------------------------------------------------------- conditional

def conditional_color_scale(
    ws: Worksheet,
    range_ref: str,
    *,
    low: str = "F8D7DA",
    mid: str = "FFF3CD",
    high: str = "D4EDDA",
) -> None:
    """3-color scale — red/yellow/green — over a given range (e.g. 'F2:F40')."""
    rule = ColorScaleRule(
        start_type="min", start_color=low,
        mid_type="percentile", mid_value=50, mid_color=mid,
        end_type="max", end_color=high,
    )
    ws.conditional_formatting.add(range_ref, rule)


# ---------------------------------------------------------------- charts

def _quoted_sheet_ref(ws: Worksheet, ref: str) -> str:
    """openpyxl requires sheet names with spaces/specials to be single-quoted
    in range strings (`'Monthly Revenue'!A1:B2`)."""
    title = ws.title
    needs_quotes = any(c in title for c in " -'!")
    if needs_quotes:
        title = "'" + title.replace("'", "''") + "'"
    return f"{title}!{ref}"


def add_bar_chart(
    ws: Worksheet,
    *,
    title: str,
    data_ref: str,
    categories_ref: str,
    anchor: str,
) -> None:
    chart = BarChart()
    chart.title = title
    chart.y_axis.title = None
    chart.x_axis.title = None
    chart.add_data(
        Reference(ws, range_string=_quoted_sheet_ref(ws, data_ref)),
        titles_from_data=True,
    )
    chart.set_categories(Reference(ws, range_string=_quoted_sheet_ref(ws, categories_ref)))
    chart.height = 9
    chart.width = 18
    ws.add_chart(chart, anchor)


def add_line_chart(
    ws: Worksheet,
    *,
    title: str,
    data_ref: str,
    categories_ref: str,
    anchor: str,
) -> None:
    chart = LineChart()
    chart.title = title
    chart.add_data(
        Reference(ws, range_string=_quoted_sheet_ref(ws, data_ref)),
        titles_from_data=True,
    )
    chart.set_categories(Reference(ws, range_string=_quoted_sheet_ref(ws, categories_ref)))
    chart.height = 9
    chart.width = 18
    ws.add_chart(chart, anchor)


# ---------------------------------------------------------------- formatting atoms

def money(value: float | None) -> float | str:
    """Return a numeric money cell or an em-dash for None (display in a money-fmt column)."""
    return value if value is not None else "—"


def pct(value: float | None) -> float | str:
    return value if value is not None else "—"


def ratio(value: float | None) -> float | str:
    return value if value is not None else "—"
