"""Sheet 9 — Board Metrics (new in Phase 6).

ARR, NRR, logo retention, expansion rate, plus MM/ENT pipeline coverage
ratios. Falls back to gap-flag cells whenever a metric's source
(BigQuery / finance mapping) is unavailable.
"""
from __future__ import annotations

from openpyxl.worksheet.worksheet import Worksheet

from agents.slt_metrics.excel_model import helpers as H, styles as S
from agents.slt_metrics.excel_model.sheets import BaseSheet
from agents.slt_metrics.pipeline.config import COVERAGE_TARGETS
from agents.slt_metrics.pipeline.planning import (
    EXPANSION_RATE_TARGET,
    LOGO_RETENTION_TARGET,
    NRR_TARGET,
)
from agents.slt_metrics.types import RevenueModelPayload


def _value_or_gap(v):
    return v if v is not None else "—"


class BoardMetricsSheet(BaseSheet):
    sheet_name = "Board Metrics"

    def write(self, ws: Worksheet, payload: RevenueModelPayload) -> None:
        bm = payload.board_metrics
        H.write_title_banner(
            ws,
            f"Board Metrics · as of {bm.as_of.isoformat()}",
            cols=3,
        )
        H.write_header_row(ws, row=2, headers=["Metric", "Value", "Target"])

        mm_target = COVERAGE_TARGETS.get("MM", 3.0)
        ent_target = COVERAGE_TARGETS.get("ENT", 4.0)
        rows: list[tuple[str, object, object, str | None]] = [
            ("ARR",                      _value_or_gap(bm.arr),                    "",                           S.FMT_MONEY),
            ("Net Revenue Retention",    _value_or_gap(bm.nrr),                    f">={NRR_TARGET:.2f}",        S.FMT_PCT),
            ("Logo Retention",           _value_or_gap(bm.logo_retention),         f">={LOGO_RETENTION_TARGET:.2f}", S.FMT_PCT),
            ("Expansion Rate",           _value_or_gap(bm.expansion_rate),         f">={EXPANSION_RATE_TARGET:.2f}", S.FMT_PCT),
            ("MM Pipeline Coverage",     _value_or_gap(bm.pipeline_coverage_mm),   f"{mm_target:g}x",            S.FMT_RATIO),
            ("ENT Pipeline Coverage",    _value_or_gap(bm.pipeline_coverage_ent),  f"{ent_target:g}x",           S.FMT_RATIO),
        ]
        for i, (label, value, target, fmt) in enumerate(rows, start=3):
            H.write_body_row(
                ws, row=i,
                values=(label, value, target),
                number_formats=(None, fmt, None),
            )

        # If unit economics fell into gap-flag, add a one-row footer note.
        if bm.unit_economics.gap_flag:
            note_row = 3 + len(rows) + 1
            ws.cell(row=note_row, column=1, value="Unit economics:")
            ws.cell(row=note_row, column=1).font = S.FONT_BODY_BOLD
            for c in range(2, 4):
                cell = ws.cell(row=note_row, column=c, value=S.GAP_TEXT)
                cell.font = S.FONT_GAP
                cell.fill = S.FILL_GAP
                cell.border = S.BORDER_CELL

        H.freeze_header(ws, rows=2)
        H.auto_width(ws)
