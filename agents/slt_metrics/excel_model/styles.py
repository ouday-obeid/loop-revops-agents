"""Shared Excel styling — palette, fonts, fills, borders, number formats.

The palette is a neutral slate-blue default (`#1F3A5F` header / `#EEF2F7`
alt-row). Swap to the Loop AI brand palette here when hex values arrive.
"""
from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


# ---------------------------------------------------------------- palette

PALETTE_HEADER_BG = "1F3A5F"        # slate blue
PALETTE_HEADER_FG = "FFFFFF"
PALETTE_ALT_ROW   = "EEF2F7"
PALETTE_BORDER    = "C3CBD6"
PALETTE_GOOD      = "2E7D32"        # ≥target coverage / attainment
PALETTE_WARN      = "B26A00"
PALETTE_BAD       = "C62828"
PALETTE_GAP_FLAG  = "8A6A00"        # amber — "Loop Pulse unavailable"
PALETTE_TITLE_BG  = "0F2540"        # header for workbook title banner


# ---------------------------------------------------------------- fills

FILL_HEADER    = PatternFill("solid", fgColor=PALETTE_HEADER_BG)
FILL_ALT_ROW   = PatternFill("solid", fgColor=PALETTE_ALT_ROW)
FILL_TITLE     = PatternFill("solid", fgColor=PALETTE_TITLE_BG)
FILL_GAP       = PatternFill("solid", fgColor="FFF4D6")


# ---------------------------------------------------------------- fonts

FONT_HEADER = Font(name="Calibri", size=11, bold=True, color=PALETTE_HEADER_FG)
FONT_TITLE  = Font(name="Calibri", size=14, bold=True, color=PALETTE_HEADER_FG)
FONT_BODY   = Font(name="Calibri", size=10)
FONT_BODY_BOLD = Font(name="Calibri", size=10, bold=True)
FONT_GAP    = Font(name="Calibri", size=10, italic=True, color=PALETTE_GAP_FLAG)


# ---------------------------------------------------------------- borders

_SIDE = Side(style="thin", color=PALETTE_BORDER)
BORDER_CELL = Border(left=_SIDE, right=_SIDE, top=_SIDE, bottom=_SIDE)


# ---------------------------------------------------------------- alignments

ALIGN_HEADER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT   = Alignment(horizontal="left", vertical="center")
ALIGN_RIGHT  = Alignment(horizontal="right", vertical="center")
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")


# ---------------------------------------------------------------- number formats

FMT_MONEY       = '"$"#,##0'
FMT_MONEY_CENTS = '"$"#,##0.00'
FMT_INT         = "#,##0"
FMT_PCT         = "0.0%"
FMT_PCT_INT     = "0%"
FMT_RATIO       = "0.00"
FMT_DATE        = "yyyy-mm-dd"

GAP_TEXT = "-- (Loop Pulse unavailable)"
