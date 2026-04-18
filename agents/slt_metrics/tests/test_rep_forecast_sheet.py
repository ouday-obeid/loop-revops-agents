"""RepForecastSheet — roster grouped by manager, QTD attainment, open pipeline."""
from __future__ import annotations

from datetime import date, datetime
from typing import Any

from openpyxl import Workbook

from agents.slt_metrics.excel_model.sheets.rep_forecast import RepForecastSheet
from agents.slt_metrics.pipeline.planning import AE_ROSTER, manager_for_ae
from agents.slt_metrics.types import (
    BoardMetrics,
    ForecastRollup,
    ForecastWeights,
    MoverSet,
    OppRecord,
    PillarScore,
    RevenueModelPayload,
    ScoredDeal,
    UnitEconomics,
)


def _closed_won(owner_name: str, acv: float) -> OppRecord:
    return OppRecord(
        id=f"cw-{owner_name}-{acv:.0f}",
        name="Closed Won",
        account_id=None, account_name=None, account_website=None, account_type=None,
        owner_id=None, owner_name=owner_name, owner_role=None, owner_manager=None,
        stage="Closed Won", is_closed=True, is_won=True,
        amount=acv, acv=acv, fixed_arr=None, locations=None,
        type="New Business", lead_source=None,
        close_date=date(2026, 4, 20),
        created_date=datetime(2026, 1, 1),
        last_activity_date=None, last_modified_date=None, last_stage_change_date=None,
        days_since_stage_change=None, time_in_stage=None, probability_sf=None,
        description=None, next_steps=None, next_step_date=None,
        icp_score=None, segment="MM",
    )


def _open_deal(owner_name: str, acv: float) -> ScoredDeal:
    return ScoredDeal(
        opp_id=f"op-{owner_name}-{acv:.0f}",
        opp_name="Open Deal",
        owner_name=owner_name,
        account_name="Acct",
        segment="MM",
        stage="Proposal",
        amount=acv, acv=acv,
        close_date=date(2026, 6, 30),
        score=65, probability=0.5, category="Commit",
        weighted_acv=acv * 0.5,
        pillars={"icp": PillarScore(0.6, "")},
        risk_flags=[],
        weights_version="v1-seed",
    )


def _payload(
    *,
    closed: list[OppRecord] | None = None,
    deals: list[ScoredDeal] | None = None,
) -> RevenueModelPayload:
    return RevenueModelPayload(
        run_date=date(2026, 4, 17),
        horizon_quarter="FY2026-Q2",
        weights=ForecastWeights(),
        scored_deals=deals or [],
        forecast_rollup=ForecastRollup(
            horizon_quarter="FY2026-Q2",
            commit_amount=0.0, best_case_amount=0.0, weighted_amount=0.0, deal_count=0,
        ),
        movers=MoverSet(period_from=date(2026, 4, 1), period_to=date(2026, 4, 17)),
        ae_cards=[], sdr_cards=[],
        board_metrics=BoardMetrics(
            as_of=date(2026, 4, 17),
            arr=None, nrr=None, logo_retention=None, expansion_rate=None,
            pipeline_coverage_mm=None, pipeline_coverage_ent=None,
            unit_economics=UnitEconomics(
                gross_revenue_retention=None, net_revenue_retention=None,
                logo_retention=None, expansion_rate=None,
                cac_payback_months=None, ltv_cac_ratio=None, gap_flag=True,
            ),
        ),
        closed_opps_quarter=closed or [],
        all_opps_snapshot=[],
    )


def _render(**kwargs) -> Any:
    wb = Workbook()
    wb.remove(wb.active)
    sheet = RepForecastSheet()
    ws = sheet.bind(wb)
    sheet.write(ws, _payload(**kwargs))
    return ws


def _find_row(ws, value: str, col: int = 1, max_rows: int = 120) -> int | None:
    for r in range(2, max_rows + 1):
        if ws.cell(row=r, column=col).value == value:
            return r
    return None


def test_rep_forecast_headers():
    ws = _render()
    assert ws.cell(row=2, column=1).value == "Rep"
    assert ws.cell(row=2, column=4).value == "Annual Quota"
    assert ws.cell(row=2, column=6).value == "QTD Attainment (of Q)"
    assert ws.cell(row=2, column=8).value == "Rep Submitted Forecast"


def test_rep_forecast_renders_manager_headers_and_members():
    ws = _render()
    # Every AE in AE_ROSTER should show up once.
    roster_names = {entry.name for entry in AE_ROSTER}
    rendered = set()
    for r in range(2, 80):
        v = ws.cell(row=r, column=1).value
        if v and v in roster_names:
            rendered.add(v)
    assert rendered == roster_names

    # Each non-empty manager group should emit a "Manager · <name>" header.
    managers_rendered = set()
    for r in range(2, 80):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str) and v.startswith("Manager · "):
            managers_rendered.add(v.removeprefix("Manager · "))
    for entry in AE_ROSTER:
        assert manager_for_ae(entry.name) in managers_rendered


def test_rep_forecast_computes_qtd_attainment_for_a_rep():
    # Sarra's annual quota = 1_000_000 → quarterly = 250_000.
    closed = [_closed_won("Sarra Herlich", 125_000.0)]
    ws = _render(closed=closed)
    sarra_row = _find_row(ws, "Sarra Herlich")
    assert sarra_row is not None
    assert ws.cell(row=sarra_row, column=4).value == 1_000_000.0
    assert ws.cell(row=sarra_row, column=5).value == 125_000.0
    # QTD attainment = 125k / 250k = 0.5.
    assert abs(ws.cell(row=sarra_row, column=6).value - 0.5) < 1e-9


def test_rep_forecast_open_pipeline_column():
    deals = [
        _open_deal("Alex Reyes", 400_000.0),
        _open_deal("Alex Reyes", 150_000.0),
    ]
    ws = _render(deals=deals)
    alex_row = _find_row(ws, "Alex Reyes")
    assert alex_row is not None
    assert ws.cell(row=alex_row, column=7).value == 550_000.0


def test_rep_forecast_submitted_forecast_column_is_blank():
    ws = _render()
    sarra_row = _find_row(ws, "Sarra Herlich")
    assert sarra_row is not None
    assert ws.cell(row=sarra_row, column=8).value is None
