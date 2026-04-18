"""Excel builder integration — synthetic payload → .xlsx → round-trip.

Verifies that:
  * every registered sheet lands in the workbook in the expected order,
  * headers are rendered on row 2 (row 1 is the title banner),
  * scored-deal data makes it into `Deal Details`,
  * the Unit Economics sheet swaps to the gap-flag banner when the BQ
    pipeline is marked unavailable.
"""
from __future__ import annotations

from datetime import date
from pathlib import Path

import pytest
from openpyxl import load_workbook

from agents.slt_metrics.excel_model import builder
from agents.slt_metrics.types import (
    AeCard,
    BoardMetrics,
    ForecastRollup,
    ForecastWeights,
    MoverSet,
    Mover,
    PillarScore,
    RevenueModelPayload,
    ScoredDeal,
    SdrCard,
    UnitEconomics,
)


EXPECTED_SHEETS = (
    "Deal Details", "AE Scorecard", "SDR Scorecard", "Unit Economics",
    "Quota", "Pipeline by Segment", "Deal Movers",
    "Forecast Summary", "Board Metrics",
    "Expansion", "Monthly Revenue", "Funnel Metrics", "Rep Forecast",
)


def _pillars() -> dict[str, PillarScore]:
    return {
        "icp":      PillarScore(0.80, "sf-icp-score"),
        "stage":    PillarScore(0.75, "Pilot"),
        "activity": PillarScore(0.60, "recent-touch"),
        "timeline": PillarScore(0.65, "in-60d"),
        "call":     PillarScore(0.50, "1-transcript"),
    }


def _deal(*, opp_id: str, score: int, weighted: float, segment: str = "ENT") -> ScoredDeal:
    return ScoredDeal(
        opp_id=opp_id,
        opp_name=f"Opp {opp_id}",
        owner_name="Jane AE",
        account_name=f"Account {opp_id}",
        segment=segment,
        stage="Pilot",
        amount=weighted * 2,
        acv=weighted * 2,
        close_date=date(2026, 5, 15),
        score=score,
        probability=score / 100.0,
        category="Strong Commit" if score >= 80 else "Commit",
        weighted_acv=weighted,
        pillars=_pillars(),
        risk_flags=["STAGE_MISMATCH"] if score < 60 else [],
        weights_version="v1-seed",
    )


def _payload(*, gap_ue: bool = False) -> RevenueModelPayload:
    scored = [
        _deal(opp_id="A1", score=85, weighted=200_000.0, segment="ENT"),
        _deal(opp_id="B2", score=55, weighted=80_000.0,  segment="MM"),
    ]
    ae_cards = [
        AeCard(
            rep_email="jane@tryloop.ai", rep_name="Jane AE",
            attainment_pct=0.62, close_rate_pct=0.40, avg_cycle_days=42.0, avg_acv=125_000.0,
            pipeline_created=500_000.0, pipeline_advanced=220_000.0,
            call_grade_avg=0.72, rep_perf_score=78,
            deals_open=6, deals_commit=2,
        ),
    ]
    sdr_cards = [
        SdrCard(
            sdr_email="sam@tryloop.ai", sdr_name="Sam SDR",
            dials=120, connects=30,
            meetings_set=12, meetings_held=8,
            pipeline_sourced=340_000.0, pipeline_advanced=90_000.0,
            leaderboard_rank=1,
        ),
    ]
    movers = MoverSet(
        period_from=date(2026, 4, 1), period_to=date(2026, 4, 8),
        movers=[
            Mover(
                opp_id="A1", opp_name="Opp A1", owner_name="Jane AE", kind="advanced",
                before={"stage": "Discovery"}, after={"stage": "Pilot"},
                delta_acv=50_000.0, delta_days=None,
            ),
        ],
    )
    rollup = ForecastRollup(
        horizon_quarter="FY2026-Q2",
        commit_amount=200_000.0,
        best_case_amount=280_000.0,
        weighted_amount=180_000.0,
        deal_count=2,
        by_owner={"jane@tryloop.ai": {"commit": 200_000.0, "best_case": 280_000.0, "weighted": 180_000.0}},
        by_segment={
            "ENT": {"commit": 200_000.0, "best_case": 200_000.0, "weighted": 170_000.0},
            "MM":  {"commit": 0.0,       "best_case": 80_000.0,  "weighted": 10_000.0},
        },
    )
    ue = (
        UnitEconomics(
            gross_revenue_retention=None, net_revenue_retention=None,
            logo_retention=None, expansion_rate=None,
            cac_payback_months=None, ltv_cac_ratio=None, gap_flag=True,
        ) if gap_ue else
        UnitEconomics(
            gross_revenue_retention=0.95, net_revenue_retention=1.12,
            logo_retention=0.92, expansion_rate=0.20,
            cac_payback_months=14.0, ltv_cac_ratio=3.8, gap_flag=False,
        )
    )
    bm = BoardMetrics(
        as_of=date(2026, 4, 1),
        arr=14_000_000.0, nrr=1.12, logo_retention=0.92, expansion_rate=0.20,
        pipeline_coverage_mm=3.2, pipeline_coverage_ent=4.5,
        unit_economics=ue,
    )
    return RevenueModelPayload(
        run_date=date(2026, 4, 13),
        horizon_quarter="FY2026-Q2",
        weights=ForecastWeights(),
        scored_deals=scored,
        forecast_rollup=rollup,
        movers=movers,
        ae_cards=ae_cards,
        sdr_cards=sdr_cards,
        board_metrics=bm,
    )


def _write_and_load(tmp_path: Path, *, gap_ue: bool = False):
    out = tmp_path / "Loop_Revenue_Model_2026-04-13.xlsx"
    returned = builder.build(_payload(gap_ue=gap_ue), out)
    assert returned == out
    assert out.exists() and out.stat().st_size > 0
    return load_workbook(out)


def test_builder_emits_all_nine_sheets_in_order(tmp_path):
    wb = _write_and_load(tmp_path)
    assert list(wb.sheetnames) == list(EXPECTED_SHEETS)


def test_builder_creates_parent_dirs(tmp_path):
    out = tmp_path / "nested" / "dir" / "Loop_Revenue_Model.xlsx"
    builder.build(_payload(), out)
    assert out.exists()


def test_deal_details_sheet_has_headers_and_row_per_deal(tmp_path):
    wb = _write_and_load(tmp_path)
    ws = wb["Deal Details"]
    # Row 1 is the banner; row 2 is headers; rows 3..N are deals.
    assert ws.cell(row=2, column=1).value == "Opp ID"
    assert ws.cell(row=2, column=10).value == "Score"
    # Two synthetic scored deals → two body rows.
    assert ws.cell(row=3, column=1).value == "A1"  # sorted by weighted_acv desc
    assert ws.cell(row=4, column=1).value == "B2"


def test_ae_scorecard_renders_body_row(tmp_path):
    wb = _write_and_load(tmp_path)
    ws = wb["AE Scorecard"]
    assert ws.cell(row=2, column=1).value == "Rep Email"
    assert ws.cell(row=3, column=1).value == "jane@tryloop.ai"
    assert ws.cell(row=3, column=11).value == 6   # Deals Open


def test_sdr_scorecard_sorts_by_rank(tmp_path):
    wb = _write_and_load(tmp_path)
    ws = wb["SDR Scorecard"]
    assert ws.cell(row=3, column=1).value == "sam@tryloop.ai"
    assert ws.cell(row=3, column=9).value == 1


def test_unit_economics_sheet_populated_when_not_gap_flagged(tmp_path):
    wb = _write_and_load(tmp_path, gap_ue=False)
    ws = wb["Unit Economics"]
    assert ws.cell(row=2, column=1).value == "Metric"
    # Six metric rows (rows 3–8).
    assert ws.cell(row=3, column=1).value == "Gross Revenue Retention"
    assert ws.cell(row=3, column=2).value == pytest.approx(0.95)
    assert ws.cell(row=3, column=3).value == "FALSE"


def test_unit_economics_sheet_emits_gap_flag_when_unavailable(tmp_path):
    wb = _write_and_load(tmp_path, gap_ue=True)
    ws = wb["Unit Economics"]
    assert ws.cell(row=3, column=2).value == "-- (Loop Pulse unavailable)"
    assert ws.cell(row=3, column=3).value == "TRUE"


def test_pipeline_by_segment_aggregates_scored_deals(tmp_path):
    wb = _write_and_load(tmp_path)
    ws = wb["Pipeline by Segment"]
    assert ws.cell(row=2, column=1).value == "Segment"
    # ENT row shows up first in canonical order.
    segments = [ws.cell(row=r, column=1).value for r in range(3, 6) if ws.cell(row=r, column=1).value]
    assert segments[0] == "ENT"
    assert "MM" in segments


def test_deal_movers_sheet_has_movers_row(tmp_path):
    wb = _write_and_load(tmp_path)
    ws = wb["Deal Movers"]
    assert ws.cell(row=2, column=1).value == "Opp ID"
    assert ws.cell(row=3, column=1).value == "A1"
    assert ws.cell(row=3, column=4).value == "advanced"


def test_forecast_summary_has_quarter_block(tmp_path):
    wb = _write_and_load(tmp_path)
    ws = wb["Forecast Summary"]
    assert ws.cell(row=2, column=1).value == "Metric"
    assert ws.cell(row=3, column=1).value == "Commit (score ≥ 80)"
    assert ws.cell(row=3, column=2).value == pytest.approx(200_000.0)


def test_board_metrics_sheet_populates_arr_and_coverage(tmp_path):
    wb = _write_and_load(tmp_path)
    ws = wb["Board Metrics"]
    assert ws.cell(row=2, column=1).value == "Metric"
    # ARR in row 3 col 2.
    assert ws.cell(row=3, column=1).value == "ARR"
    assert ws.cell(row=3, column=2).value == pytest.approx(14_000_000.0)


def test_board_metrics_shows_ue_footer_when_gap_flagged(tmp_path):
    wb = _write_and_load(tmp_path, gap_ue=True)
    ws = wb["Board Metrics"]
    # Six metric rows + header row = 8 filled rows; footer note lives below.
    for row in range(2, 15):
        if ws.cell(row=row, column=1).value == "Unit economics:":
            assert ws.cell(row=row, column=2).value == "-- (Loop Pulse unavailable)"
            break
    else:
        pytest.fail("Expected 'Unit economics:' gap footer row in Board Metrics sheet")


def test_build_rejects_empty_sheet_registry(tmp_path):
    out = tmp_path / "empty.xlsx"
    with pytest.raises(ValueError, match="at least one sheet"):
        builder.build(_payload(), out, sheets=[])
