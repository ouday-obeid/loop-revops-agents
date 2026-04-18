"""FunnelMetricsSheet — stage distribution + segment close-rate blocks."""
from __future__ import annotations

from datetime import date, datetime
from typing import Any

from openpyxl import Workbook

from agents.slt_metrics.excel_model.sheets.funnel_metrics import FunnelMetricsSheet
from agents.slt_metrics.pipeline.planning import quarterly_funnel_target
from agents.slt_metrics.types import (
    BoardMetrics,
    ForecastRollup,
    ForecastWeights,
    MoverSet,
    OppRecord,
    RevenueModelPayload,
    UnitEconomics,
)


def _opp(*, id_: str, stage: str, acv: float, is_closed: bool = False,
         is_won: bool = False, segment: str = "MM") -> OppRecord:
    return OppRecord(
        id=id_, name=id_,
        account_id=None, account_name=None, account_website=None, account_type=None,
        owner_id=None, owner_name=None, owner_role=None, owner_manager=None,
        stage=stage, is_closed=is_closed, is_won=is_won,
        amount=acv, acv=acv, fixed_arr=None, locations=None,
        type="New Business", lead_source=None,
        close_date=date(2026, 4, 20),
        created_date=datetime(2026, 1, 1),
        last_activity_date=None, last_modified_date=None, last_stage_change_date=None,
        days_since_stage_change=None, time_in_stage=None, probability_sf=None,
        description=None, next_steps=None, next_step_date=None,
        icp_score=None, segment=segment,
    )


def _payload(
    *,
    all_opps: list[OppRecord] | None = None,
    closed: list[OppRecord] | None = None,
) -> RevenueModelPayload:
    return RevenueModelPayload(
        run_date=date(2026, 4, 17),
        horizon_quarter="FY2026-Q2",
        weights=ForecastWeights(),
        scored_deals=[],
        forecast_rollup=ForecastRollup(
            horizon_quarter="FY2026-Q2",
            commit_amount=0.0, best_case_amount=0.0, weighted_amount=0.0, deal_count=0,
        ),
        movers=MoverSet(period_from=date(2026, 4, 1), period_to=date(2026, 4, 17)),
        ae_cards=[], sdr_cards=[],
        board_metrics=BoardMetrics(
            as_of=date(2026, 4, 17),
            arr=None, nrr=None, logo_retention=None, expansion_rate=None,
            pipeline_coverage_mm=None, pipeline_coverage_ent=None,
            unit_economics=UnitEconomics(
                gross_revenue_retention=None, net_revenue_retention=None,
                logo_retention=None, expansion_rate=None,
                cac_payback_months=None, ltv_cac_ratio=None, gap_flag=True,
            ),
        ),
        closed_opps_quarter=closed or [],
        all_opps_snapshot=all_opps or [],
    )


def _render(**kwargs) -> Any:
    wb = Workbook()
    wb.remove(wb.active)
    sheet = FunnelMetricsSheet()
    ws = sheet.bind(wb)
    sheet.write(ws, _payload(**kwargs))
    return ws


def test_funnel_stage_block_in_canonical_order():
    all_opps = [
        _opp(id_="p1", stage="Proposal", acv=100_000.0),
        _opp(id_="p2", stage="Proposal", acv=50_000.0),
        _opp(id_="d1", stage="Demo",     acv=25_000.0),
        _opp(id_="nms", stage="New Meeting Set", acv=10_000.0),
    ]
    ws = _render(all_opps=all_opps)
    # Canonical STAGES order is "No Show", "Disqualified", "Closed Lost",
    # "New Meeting Set", "Demo", "Business Case", "Pilot", "Proposal", "Closed Won".
    # Only the stages present should appear.
    assert ws.cell(row=2, column=1).value == "Stage"
    assert ws.cell(row=3, column=1).value == "New Meeting Set"
    assert ws.cell(row=4, column=1).value == "Demo"
    assert ws.cell(row=5, column=1).value == "Proposal"
    # Proposal count = 2, ACV = 150k.
    assert ws.cell(row=5, column=2).value == 2
    assert ws.cell(row=5, column=3).value == 150_000.0


def test_funnel_stage_block_excludes_closed_opps():
    all_opps = [
        _opp(id_="open", stage="Demo",       acv=50_000.0, is_closed=False),
        _opp(id_="won",  stage="Closed Won", acv=500_000.0, is_closed=True, is_won=True),
    ]
    ws = _render(all_opps=all_opps)
    # Only the open opp's stage appears.
    first_stage = ws.cell(row=3, column=1).value
    assert first_stage == "Demo"
    # No Closed Won row in the stage block.
    for r in range(3, 6):
        assert ws.cell(row=r, column=1).value != "Closed Won"


def test_funnel_segment_block_renders_targets_and_coverage():
    closed = [
        _opp(id_="w1", stage="Closed Won",  acv=100_000.0, is_closed=True, is_won=True, segment="MM"),
        _opp(id_="w2", stage="Closed Won",  acv=90_000.0,  is_closed=True, is_won=True, segment="MM"),
        _opp(id_="l1", stage="Closed Lost", acv=50_000.0,  is_closed=True, is_won=False, segment="MM"),
        _opp(id_="w3", stage="Closed Won",  acv=300_000.0, is_closed=True, is_won=True, segment="ENT"),
    ]
    ws = _render(closed=closed)
    # Locate the segment block header (after stage block + gap row).
    header_row = None
    for r in range(2, 20):
        if ws.cell(row=r, column=1).value == "Segment":
            header_row = r
            break
    assert header_row is not None

    # Row layout: ENT, MM, SMB.
    ent_row = header_row + 1
    mm_row = header_row + 2
    smb_row = header_row + 3

    assert ws.cell(row=ent_row, column=1).value == "ENT"
    assert ws.cell(row=mm_row,  column=1).value == "MM"
    assert ws.cell(row=smb_row, column=1).value == "SMB"

    # MM: 2 won / 1 lost, win rate = 2/3.
    assert ws.cell(row=mm_row, column=2).value == 2
    assert ws.cell(row=mm_row, column=3).value == 190_000.0
    assert ws.cell(row=mm_row, column=4).value == 1
    assert abs(ws.cell(row=mm_row, column=6).value - (2 / 3)) < 1e-9

    # Target pulled from QUARTERLY_FUNNEL_TARGETS["Q2"]["MM"].
    assert ws.cell(row=mm_row, column=7).value == quarterly_funnel_target("Q2", "MM")
    # Closed count column = won + lost = 3.
    assert ws.cell(row=mm_row, column=8).value == 3


def test_funnel_segment_block_zero_targets_zero_coverage():
    ws = _render(closed=[])
    # Find header row.
    header_row = None
    for r in range(2, 20):
        if ws.cell(row=r, column=1).value == "Segment":
            header_row = r
            break
    assert header_row is not None
    # ENT row with no data.
    ent_row = header_row + 1
    assert ws.cell(row=ent_row, column=2).value == 0  # won_count
    assert ws.cell(row=ent_row, column=9).value == 0.0  # coverage
