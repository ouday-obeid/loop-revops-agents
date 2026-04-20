"""MonthlyRevenueSheet — target+actual for new_biz, expansion, total."""
from __future__ import annotations

from datetime import date, datetime
from typing import Any

from openpyxl import Workbook

from agents.slt_metrics.excel_model.sheets.monthly_revenue import MonthlyRevenueSheet
from agents.slt_metrics.pipeline.planning import MONTHLY_TARGETS
from agents.slt_metrics.types import (
    BoardMetrics,
    ForecastRollup,
    ForecastWeights,
    MoverSet,
    OppRecord,
    RevenueModelPayload,
    UnitEconomics,
)


def _closed_won(*, kind_type: str, month: int, acv: float) -> OppRecord:
    return OppRecord(
        id=f"cw-{kind_type}-{month}-{acv:.0f}",
        name="Closed Won",
        account_id=None, account_name=None, account_website=None, account_type=None,
        owner_id=None, owner_name=None, owner_role=None, owner_manager=None,
        stage="Closed Won", is_closed=True, is_won=True,
        amount=acv, acv=acv, fixed_arr=None, locations=None,
        type=kind_type, lead_source=None,
        close_date=date(2026, month, 15),
        created_date=datetime(2026, 1, 1),
        last_activity_date=None, last_modified_date=None, last_stage_change_date=None,
        days_since_stage_change=None, time_in_stage=None, probability_sf=None,
        description=None, next_steps=None, next_step_date=None,
        icp_score=None, segment="MM",
    )


def _payload(closed: list[OppRecord]) -> RevenueModelPayload:
    return RevenueModelPayload(
        run_date=date(2026, 4, 17),
        horizon_quarter="FY2026-Q2",
        weights=ForecastWeights(),
        scored_deals=[],
        forecast_rollup=ForecastRollup(
            horizon_quarter="FY2026-Q2",
            commit_amount=0.0, best_case_amount=0.0, weighted_amount=0.0, deal_count=0,
        ),
        movers=MoverSet(period_from=date(2026, 4, 1), period_to=date(2026, 4, 17)),
        ae_cards=[], sdr_cards=[],
        board_metrics=BoardMetrics(
            as_of=date(2026, 4, 17),
            arr=None, nrr=None, logo_retention=None, expansion_rate=None,
            pipeline_coverage_mm=None, pipeline_coverage_ent=None,
            unit_economics=UnitEconomics(
                gross_revenue_retention=None, net_revenue_retention=None,
                logo_retention=None, expansion_rate=None,
                cac_payback_months=None, ltv_cac_ratio=None, gap_flag=True,
            ),
        ),
        closed_opps_quarter=closed,
    )


def _render(closed: list[OppRecord]) -> Any:
    wb = Workbook()
    wb.remove(wb.active)
    sheet = MonthlyRevenueSheet()
    ws = sheet.bind(wb)
    sheet.write(ws, _payload(closed))
    return ws


def test_monthly_revenue_headers_and_12_month_rows():
    ws = _render([])
    assert ws.cell(row=2, column=1).value == "Month"
    assert ws.cell(row=2, column=2).value == "Target New Biz"
    assert ws.cell(row=2, column=10).value == "Total Attainment"
    assert ws.cell(row=3, column=1).value == "Jan"
    assert ws.cell(row=14, column=1).value == "Dec"
    assert ws.cell(row=15, column=1).value == "Total"


def test_monthly_revenue_targets_from_planning():
    ws = _render([])
    assert ws.cell(row=3, column=2).value == MONTHLY_TARGETS[1].new_biz
    assert ws.cell(row=3, column=5).value == MONTHLY_TARGETS[1].expansion
    # Target Total = new_biz + expansion.
    assert ws.cell(row=3, column=8).value == (
        MONTHLY_TARGETS[1].new_biz + MONTHLY_TARGETS[1].expansion
    )


def test_monthly_revenue_splits_actual_by_kind():
    closed = [
        _closed_won(kind_type="New Business", month=4, acv=500_000.0),
        _closed_won(kind_type="Expansion",    month=4, acv=75_000.0),
    ]
    ws = _render(closed)
    # Row 6 = April.
    assert ws.cell(row=6, column=3).value == 500_000.0   # Actual New Biz
    assert ws.cell(row=6, column=6).value == 75_000.0    # Actual Expansion
    assert ws.cell(row=6, column=9).value == 575_000.0   # Actual Total


def test_monthly_revenue_attainment_columns():
    closed = [
        _closed_won(
            kind_type="New Business", month=4,
            acv=MONTHLY_TARGETS[4].new_biz,
        ),
    ]
    ws = _render(closed)
    assert ws.cell(row=6, column=4).value == 1.0   # New Biz Attainment
    assert ws.cell(row=6, column=7).value == 0.0   # Expansion Attainment (no expansion)


def test_monthly_revenue_total_row_sums_year():
    ws = _render([])
    expected_target_nb = sum(MONTHLY_TARGETS[m].new_biz for m in range(1, 13))
    expected_target_ex = sum(MONTHLY_TARGETS[m].expansion for m in range(1, 13))
    assert ws.cell(row=15, column=2).value == expected_target_nb
    assert ws.cell(row=15, column=5).value == expected_target_ex
    assert ws.cell(row=15, column=8).value == expected_target_nb + expected_target_ex
