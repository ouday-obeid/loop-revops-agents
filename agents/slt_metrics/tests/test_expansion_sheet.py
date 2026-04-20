"""ExpansionSheet — monthly target vs actual rendering."""
from __future__ import annotations

from datetime import date, datetime
from typing import Any

from openpyxl import Workbook

from agents.slt_metrics.excel_model.sheets.expansion import ExpansionSheet
from agents.slt_metrics.pipeline.planning import MONTHLY_TARGETS
from agents.slt_metrics.types import (
    BoardMetrics,
    ForecastRollup,
    ForecastWeights,
    MoverSet,
    OppRecord,
    RevenueModelPayload,
    UnitEconomics,
)


def _closed_won_opp(*, kind_type: str, month: int, acv: float) -> OppRecord:
    return OppRecord(
        id=f"cw-{kind_type}-{month}-{acv:.0f}",
        name="Closed Won",
        account_id=None, account_name=None, account_website=None, account_type=None,
        owner_id=None, owner_name=None, owner_role=None, owner_manager=None,
        stage="Closed Won", is_closed=True, is_won=True,
        amount=acv, acv=acv, fixed_arr=None, locations=None,
        type=kind_type, lead_source=None,
        close_date=date(2026, month, 15),
        created_date=datetime(2026, 1, 1),
        last_activity_date=None, last_modified_date=None, last_stage_change_date=None,
        days_since_stage_change=None, time_in_stage=None, probability_sf=None,
        description=None, next_steps=None, next_step_date=None,
        icp_score=None, segment="MM",
    )


def _payload(closed: list[OppRecord]) -> RevenueModelPayload:
    return RevenueModelPayload(
        run_date=date(2026, 4, 17),
        horizon_quarter="FY2026-Q2",
        weights=ForecastWeights(),
        scored_deals=[],
        forecast_rollup=ForecastRollup(
            horizon_quarter="FY2026-Q2",
            commit_amount=0.0, best_case_amount=0.0, weighted_amount=0.0, deal_count=0,
        ),
        movers=MoverSet(period_from=date(2026, 4, 1), period_to=date(2026, 4, 17)),
        ae_cards=[], sdr_cards=[],
        board_metrics=BoardMetrics(
            as_of=date(2026, 4, 17),
            arr=None, nrr=None, logo_retention=None, expansion_rate=None,
            pipeline_coverage_mm=None, pipeline_coverage_ent=None,
            unit_economics=UnitEconomics(
                gross_revenue_retention=None, net_revenue_retention=None,
                logo_retention=None, expansion_rate=None,
                cac_payback_months=None, ltv_cac_ratio=None, gap_flag=True,
            ),
        ),
        closed_opps_quarter=closed,
    )


def _render(closed: list[OppRecord]) -> Any:
    wb = Workbook()
    wb.remove(wb.active)
    sheet = ExpansionSheet()
    ws = sheet.bind(wb)
    sheet.write(ws, _payload(closed))
    return ws


def test_expansion_sheet_headers_and_row_count():
    ws = _render([])
    assert ws.cell(row=2, column=1).value == "Month"
    assert ws.cell(row=2, column=2).value == "Target Expansion"
    assert ws.cell(row=2, column=5).value == "Attainment"
    # 12 month rows (3..14) + Total row (15) = 13 body rows.
    assert ws.cell(row=3, column=1).value == "Jan"
    assert ws.cell(row=14, column=1).value == "Dec"
    assert ws.cell(row=15, column=1).value == "Total"


def test_expansion_sheet_renders_targets_from_planning():
    ws = _render([])
    # Row 3 = January, column 2 = Target Expansion.
    assert ws.cell(row=3, column=2).value == MONTHLY_TARGETS[1].expansion
    assert ws.cell(row=4, column=2).value == MONTHLY_TARGETS[2].expansion
    # Empty closed list → all actuals 0, Δ = -target.
    assert ws.cell(row=3, column=3).value == 0.0
    assert ws.cell(row=3, column=4).value == -MONTHLY_TARGETS[1].expansion


def test_expansion_sheet_sums_actuals_by_month():
    closed = [
        _closed_won_opp(kind_type="Expansion", month=4, acv=60_000.0),
        _closed_won_opp(kind_type="Upsell",    month=4, acv=40_000.0),
        _closed_won_opp(kind_type="New Business", month=4, acv=200_000.0),  # excluded
        _closed_won_opp(kind_type="Expansion", month=5, acv=25_000.0),
    ]
    ws = _render(closed)
    # Row 6 = April, row 7 = May.
    assert ws.cell(row=6, column=3).value == 100_000.0
    assert ws.cell(row=7, column=3).value == 25_000.0
    # Total row should sum actuals: 100k + 25k = 125k.
    assert ws.cell(row=15, column=3).value == 125_000.0


def test_expansion_sheet_attainment_column_is_ratio():
    closed = [
        _closed_won_opp(kind_type="Expansion", month=4, acv=MONTHLY_TARGETS[4].expansion),
    ]
    ws = _render(closed)
    # April attainment = actual / target = 1.0.
    assert ws.cell(row=6, column=5).value == 1.0


def test_expansion_sheet_total_target_sums_all_months():
    ws = _render([])
    expected_total_target = sum(MONTHLY_TARGETS[m].expansion for m in range(1, 13))
    assert ws.cell(row=15, column=2).value == expected_total_target
