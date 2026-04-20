"""Team Deal Breakdown sheets — one tab per manager team with deal-level detail."""

from __future__ import annotations

from openpyxl.worksheet.worksheet import Worksheet

from sheets.base import BaseSheet
from formatting.styles import FONT_BODY_BOLD, THIN_BORDER
from formatting.utils import auto_width, freeze_panes


# Columns for deal-level rows: (df_column, display_name, format_type)
DEAL_COLUMNS = [
    ("owner", "Rep", None),
    ("organization", "Organization", None),
    ("opp_name", "Opportunity", None),
    ("stage", "Stage", None),
    ("acv", "ACV", "currency"),
    ("locations", "Locations", "number"),
    ("weighted_acv", "Weighted ACV", "currency"),
    ("segment", "Segment", None),
    ("close_date", "Close Date", None),
    ("age", "Age (Days)", "number"),
    ("lead_source", "Lead Source", None),
    ("products", "Products", None),
    ("contract_type", "Contract Type", None),
    ("opp_notes", "Notes", None),
]

# Extra columns appended when forecast data is available
FORECAST_COLUMNS = [
    ("Forecast Tier", None),
    ("Forecast ACV", "currency"),
    ("ACV Delta", "currency"),
]

# Column indices (1-based) for numeric subtotals in base columns
COL_ACV = 5
COL_LOCS = 6
COL_WEIGHTED = 7


class TeamDealsSheet(BaseSheet):
    """Creates one sheet per manager team with deal-level breakdowns by rep."""

    sheet_name = "Team Deals"  # placeholder — actual tabs named per team

    def generate(self) -> Worksheet | None:
        """Override to create multiple sheets — one per team."""
        if not self._should_generate():
            return None

        last_ws = None
        for mgr, members in self.cfg.manager_groups.items():
            tab_name = f"{mgr} Deals"[:31]  # Excel max 31 chars
            ws = self.wb.create_sheet(title=tab_name)
            self._write_team(ws, mgr, members)
            self._format(ws)
            last_ws = ws

        return last_ws

    def _should_generate(self) -> bool:
        return bool(self.cfg.manager_groups)

    def _write(self, ws: Worksheet) -> None:
        pass  # Not used — see _write_team

    def _build_forecast_lookup(self, members: list[str]) -> dict:
        """Build deal name → (tier, acv) lookup from forecast data for team members.

        Matches forecast deals to SF opps by normalized organization name.
        """
        if self.proc.rep_forecast is None:
            return {}

        matched_deals = self.proc.rep_forecast.get("matched_deals", {})
        lookup = {}  # sf_opp_name -> (tier, forecast_acv, acv_delta)

        for rep_name in members:
            rep_matched = matched_deals.get(rep_name, [])
            for md in rep_matched:
                if md.matched and md.sf_opp_name:
                    lookup[md.sf_opp_name] = (
                        md.forecast_deal.tier,
                        md.forecast_deal.acv,
                        md.acv_delta,
                    )
                # Also match by organization name
                if md.matched and md.sf_org:
                    lookup[md.sf_org] = (
                        md.forecast_deal.tier,
                        md.forecast_deal.acv,
                        md.acv_delta,
                    )

        return lookup

    def _write_team(self, ws: Worksheet, mgr: str, members: list[str]) -> None:
        """Write deal-by-deal breakdown for one team, grouped by rep."""
        pipe = self.proc.forecastable_pipeline
        team_deals = pipe[pipe["owner"].isin(members)].copy()
        team_deals = team_deals.sort_values(["owner", "acv"], ascending=[True, False])

        has_forecast = self.proc.rep_forecast is not None
        forecast_lookup = self._build_forecast_lookup(members)

        # Determine total column count
        all_headers = [c[1] for c in DEAL_COLUMNS]
        if has_forecast:
            all_headers.extend([c[0] for c in FORECAST_COLUMNS])
        num_cols = len(all_headers)

        total_acv = float(team_deals["acv"].sum())
        total_weighted = float(team_deals["weighted_acv"].sum())
        total_locs = int(team_deals["locations"].sum()) if "locations" in team_deals.columns else 0
        total_deals = len(team_deals)

        # Row 1 — title
        self._write_section_title(ws, 1, f"{mgr} Team — Pipeline Deals to Forecast")

        # Row 2 — summary
        summary = (f"{total_deals} deals  |  ACV: ${total_acv:,.0f}  |  "
                   f"Weighted: ${total_weighted:,.0f}  |  Locations: {total_locs:,}")
        if has_forecast:
            # Count how many deals have forecast matches
            matched_count = 0
            for _, deal in team_deals.iterrows():
                opp = str(deal.get("opp_name", ""))
                org = str(deal.get("organization", ""))
                if opp in forecast_lookup or org in forecast_lookup:
                    matched_count += 1
            summary += f"  |  Forecasted: {matched_count}/{total_deals}"
        self._write_cell(ws, 2, 1, summary, bold=True)

        # Row 4 — column headers
        self._write_headers(ws, 4, all_headers)

        row = 5
        current_rep = None
        rep_acv = 0.0
        rep_locs = 0
        rep_weighted = 0.0
        rep_count = 0

        for _, deal in team_deals.iterrows():
            rep_name = str(deal.get("owner", ""))

            # When rep changes, write subtotal for previous rep
            if rep_name != current_rep:
                if current_rep is not None and rep_count > 0:
                    row = self._write_subtotal(
                        ws, row, current_rep, rep_count, rep_acv, rep_locs, rep_weighted, num_cols
                    )
                    row += 1  # blank separator row

                current_rep = rep_name
                rep_acv = 0.0
                rep_locs = 0
                rep_weighted = 0.0
                rep_count = 0

            # Write base deal columns
            for c_idx, (col_name, _, fmt) in enumerate(DEAL_COLUMNS, start=1):
                val = deal.get(col_name, "")
                if hasattr(val, "item"):
                    val = val.item()
                if str(val) in ("nan", "NaT"):
                    val = ""
                self._write_cell(ws, row, c_idx, val, fmt=fmt)

            # Write forecast columns if available
            if has_forecast:
                fc_start = len(DEAL_COLUMNS) + 1
                opp = str(deal.get("opp_name", ""))
                org = str(deal.get("organization", ""))
                match = forecast_lookup.get(opp) or forecast_lookup.get(org)
                if match:
                    tier, fc_acv, delta = match
                    self._write_cell(ws, row, fc_start, tier)
                    self._write_cell(ws, row, fc_start + 1, fc_acv, fmt="currency")
                    self._write_cell(ws, row, fc_start + 2, delta, fmt="currency")

            acv_val = deal.get("acv", 0)
            loc_val = deal.get("locations", 0)
            wtd_val = deal.get("weighted_acv", 0)
            rep_acv += float(acv_val) if str(acv_val) != "nan" else 0
            rep_locs += int(loc_val) if str(loc_val) != "nan" else 0
            rep_weighted += float(wtd_val) if str(wtd_val) != "nan" else 0
            rep_count += 1
            row += 1

        # Final rep subtotal
        if current_rep is not None and rep_count > 0:
            row = self._write_subtotal(
                ws, row, current_rep, rep_count, rep_acv, rep_locs, rep_weighted, num_cols
            )

        # Team total
        row += 1
        self._write_cell(ws, row, 1, f"TEAM TOTAL — {mgr}", bold=True)
        self._write_cell(ws, row, 4, f"{total_deals} deals", bold=True)
        self._write_cell(ws, row, COL_ACV, total_acv, fmt="currency", bold=True)
        self._write_cell(ws, row, COL_LOCS, total_locs, fmt="number", bold=True)
        self._write_cell(ws, row, COL_WEIGHTED, total_weighted, fmt="currency", bold=True)
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.border = THIN_BORDER
            cell.font = FONT_BODY_BOLD

        # Auto-filter on headers
        if ws.max_row > 4:
            from openpyxl.utils import get_column_letter
            last_col = get_column_letter(num_cols)
            ws.auto_filter.ref = f"A4:{last_col}{ws.max_row}"

    def _write_subtotal(self, ws, row, rep_name, count, acv, locs, weighted, num_cols):
        """Write a rep subtotal row and return the next row number."""
        self._write_cell(ws, row, 1, f"  {rep_name} Subtotal ({count} deals)", bold=True)
        self._write_cell(ws, row, COL_ACV, acv, fmt="currency", bold=True)
        self._write_cell(ws, row, COL_LOCS, locs, fmt="number", bold=True)
        self._write_cell(ws, row, COL_WEIGHTED, weighted, fmt="currency", bold=True)
        for c in range(1, num_cols + 1):
            ws.cell(row=row, column=c).border = THIN_BORDER
        return row + 1

    def _format(self, ws: Worksheet) -> None:
        auto_width(ws, min_width=10, max_width=40)
        freeze_panes(ws, row=5, col=2)
        self._alt_row_shading(ws, 5, ws.max_row)
