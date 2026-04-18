"""Abstract base sheet writer with shared formatting helpers."""

from __future__ import annotations

from abc import ABC, abstractmethod

from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from core.config_schema import Config
from core.processor import Processor
from formatting.styles import (
    ALIGN_CENTER, ALIGN_RIGHT, FILL_ALT_ROW, FILL_HEADER, FILL_SUBHEADER,
    FONT_BODY, FONT_BODY_BOLD, FONT_HEADER, FONT_SUBHEADER, FONT_TITLE,
    FMT_CURRENCY, FMT_NUMBER, FMT_PERCENT, THIN_BORDER,
)
from formatting.utils import auto_width, freeze_panes, write_header_row, write_title


class BaseSheet(ABC):
    """Base class for all sheet writers."""

    sheet_name: str = "Sheet"

    def __init__(self, wb: Workbook, cfg: Config, proc: Processor):
        self.wb = wb
        self.cfg = cfg
        self.proc = proc

    def generate(self) -> Worksheet | None:
        """Create the sheet, write content, apply formatting."""
        if not self._should_generate():
            return None
        ws = self.wb.create_sheet(title=self.sheet_name)
        self._write(ws)
        self._format(ws)
        return ws

    def _should_generate(self) -> bool:
        """Override to conditionally skip sheet generation."""
        return True

    @abstractmethod
    def _write(self, ws: Worksheet) -> None:
        """Write all data to the sheet. Subclasses implement this."""

    def _format(self, ws: Worksheet) -> None:
        """Apply default formatting. Override for custom behavior."""
        auto_width(ws)
        if ws.max_row > 1:
            freeze_panes(ws, row=2, col=1)

    # --- Convenience helpers ---

    def _write_section_title(self, ws: Worksheet, row: int, title: str, col: int = 1):
        write_title(ws, row, title, col)

    def _write_headers(self, ws: Worksheet, row: int, headers: list[str], start_col: int = 1):
        write_header_row(ws, row, headers, start_col)

    def _write_subheader(self, ws: Worksheet, row: int, headers: list[str], start_col: int = 1):
        for i, h in enumerate(headers):
            cell = ws.cell(row=row, column=start_col + i, value=h)
            cell.font = FONT_SUBHEADER
            cell.fill = FILL_SUBHEADER
            cell.alignment = ALIGN_CENTER
            cell.border = THIN_BORDER

    def _write_cell(self, ws, row, col, value, fmt=None, bold=False):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = FONT_BODY_BOLD if bold else FONT_BODY
        if fmt == "currency":
            cell.number_format = FMT_CURRENCY
            cell.alignment = ALIGN_RIGHT
        elif fmt == "percent":
            cell.number_format = FMT_PERCENT
            cell.alignment = ALIGN_RIGHT
        elif fmt == "number":
            cell.number_format = FMT_NUMBER
            cell.alignment = ALIGN_RIGHT
        return cell

    def _alt_row_shading(self, ws: Worksheet, start_row: int, end_row: int,
                         start_col: int = 1, end_col: int | None = None):
        if end_col is None:
            end_col = ws.max_column
        for r in range(start_row, end_row + 1):
            if (r - start_row) % 2 == 1:
                for c in range(start_col, end_col + 1):
                    ws.cell(row=r, column=c).fill = FILL_ALT_ROW
