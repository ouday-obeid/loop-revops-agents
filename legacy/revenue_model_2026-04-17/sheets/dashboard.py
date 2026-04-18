"""Dashboard - Executive summary with KPI cards."""

from __future__ import annotations

from openpyxl.chart import Reference
from openpyxl.worksheet.worksheet import Worksheet

from core.processor import MONTH_NAMES, MONTHS
from formatting.charts import create_bar_chart, create_pie_chart
from formatting.styles import (
    ALIGN_CENTER, FILL_KPI_BG, FONT_KPI_LABEL, FONT_KPI_VALUE, FONT_TITLE,
    FMT_CURRENCY, FMT_PERCENT, FMT_NUMBER, FILL_HEADER, FONT_HEADER,
    THIN_BORDER,
)
from formatting.utils import auto_width, merge_and_center
from sheets.base import BaseSheet


class DashboardSheet(BaseSheet):
    sheet_name = "Dashboard"

    def _write(self, ws: Worksheet) -> None:
        proc = self.proc
        cfg = self.cfg

        # === Title ===
        cell = merge_and_center(ws, 1, 1, 1, 10, "2026 Revenue Model - Executive Dashboard")
        cell.font = FONT_TITLE

        # === KPI Cards (Row 3-5) ===
        kpis = [
            ("YTD Closed Won", proc.ytd_closed_won_acv, "currency"),
            ("YTD NB Revenue", proc.ytd_closed_won_nb_acv, "currency"),
            ("YTD Expansion", proc.ytd_closed_won_exp_acv, "currency"),
            ("NB Attainment", proc.ytd_attainment_nb, "percent"),
            ("Open Pipeline", proc.total_pipeline_acv, "currency"),
            ("Weighted Pipeline", proc.total_weighted_pipeline, "currency"),
            ("Pipeline Coverage", proc.pipeline_coverage, "number"),
            ("Win Rate", proc.overall_win_rate, "percent"),
            ("Deals Won", proc.ytd_closed_won_count, "number"),
            ("Avg Deal Size", proc.avg_deal_size, "currency"),
        ]

        for i, (label, value, fmt) in enumerate(kpis):
            col = 1 + (i * 2)
            if col + 1 > 20:
                break
            # KPI label
            lcell = ws.cell(row=3, column=col, value=label)
            lcell.font = FONT_KPI_LABEL
            lcell.fill = FILL_KPI_BG
            lcell.alignment = ALIGN_CENTER
            ws.cell(row=3, column=col + 1).fill = FILL_KPI_BG
            ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col + 1)

            # KPI value
            vcell = ws.cell(row=4, column=col, value=value)
            vcell.font = FONT_KPI_VALUE
            vcell.alignment = ALIGN_CENTER
            ws.cell(row=4, column=col + 1).fill = FILL_KPI_BG
            ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col + 1)
            if fmt == "currency":
                vcell.number_format = FMT_CURRENCY
            elif fmt == "percent":
                vcell.number_format = FMT_PERCENT
            elif fmt == "number":
                vcell.number_format = FMT_NUMBER

        # === Annual Targets Summary (Row 7) ===
        self._write_section_title(ws, 7, "Annual Targets")
        target_headers = ["Metric", "Target", "YTD Actual", "Remaining", "Attainment %"]
        self._write_headers(ws, 8, target_headers)

        # NB row
        nb_target = cfg.targets["net_new_arr"]
        nb_actual = proc.ytd_closed_won_nb_acv
        self._write_cell(ws, 9, 1, "Net New ARR", bold=True)
        self._write_cell(ws, 9, 2, nb_target, fmt="currency")
        self._write_cell(ws, 9, 3, nb_actual, fmt="currency")
        self._write_cell(ws, 9, 4, nb_target - nb_actual, fmt="currency")
        self._write_cell(ws, 9, 5, nb_actual / nb_target if nb_target else 0, fmt="percent")

        # Expansion row
        exp_target = cfg.targets["expansion_arr"]
        exp_actual = proc.ytd_closed_won_exp_acv
        self._write_cell(ws, 10, 1, "Expansion ARR", bold=True)
        self._write_cell(ws, 10, 2, exp_target, fmt="currency")
        self._write_cell(ws, 10, 3, exp_actual, fmt="currency")
        self._write_cell(ws, 10, 4, exp_target - exp_actual, fmt="currency")
        self._write_cell(ws, 10, 5, exp_actual / exp_target if exp_target else 0, fmt="percent")

        # Total row
        total_target = cfg.targets["gross_new_arr"]
        total_actual = proc.ytd_closed_won_acv
        self._write_cell(ws, 11, 1, "Gross New ARR", bold=True)
        self._write_cell(ws, 11, 2, total_target, fmt="currency")
        self._write_cell(ws, 11, 3, total_actual, fmt="currency")
        self._write_cell(ws, 11, 4, total_target - total_actual, fmt="currency")
        self._write_cell(ws, 11, 5, total_actual / total_target if total_target else 0, fmt="percent")

        # === Unit Economics: ARPL, ADS, LPD (Row 14) ===
        self._write_section_title(ws, 14, "Unit Economics - Target vs Current")
        ue_headers = ["Metric", "SMB Target", "SMB Current", "MM Target", "MM Current",
                      "Ent Target", "Ent Current", "Blended Target", "Blended Current"]
        self._write_headers(ws, 15, ue_headers)

        # ARPL row
        self._write_cell(ws, 16, 1, "ARPL", bold=True)
        for i, seg in enumerate(["SMB", "MM", "Ent"]):
            self._write_cell(ws, 16, 2 + i * 2, cfg.segment_target_arpl(seg), fmt="currency")
            self._write_cell(ws, 16, 3 + i * 2, proc.current_arpl.get(seg, 0), fmt="currency")
        self._write_cell(ws, 16, 8, cfg.blended_targets.get("arpl", 0), fmt="currency")
        self._write_cell(ws, 16, 9, proc.current_arpl.get("Blended", 0), fmt="currency")

        # ADS row
        self._write_cell(ws, 17, 1, "ADS", bold=True)
        for i, seg in enumerate(["SMB", "MM", "Ent"]):
            self._write_cell(ws, 17, 2 + i * 2, cfg.segment_target_ads(seg), fmt="currency")
            self._write_cell(ws, 17, 3 + i * 2, proc.current_ads.get(seg, 0), fmt="currency")
        self._write_cell(ws, 17, 8, cfg.blended_targets.get("ads", 0), fmt="currency")
        self._write_cell(ws, 17, 9, proc.current_ads.get("Blended", 0), fmt="currency")

        # LPD row
        self._write_cell(ws, 18, 1, "Avg Locs/Deal", bold=True)
        for i, seg in enumerate(["SMB", "MM", "Ent"]):
            self._write_cell(ws, 18, 2 + i * 2, cfg.segment_target_lpd(seg), fmt="number")
            self._write_cell(ws, 18, 3 + i * 2, proc.current_lpd.get(seg, 0), fmt="number")
        self._write_cell(ws, 18, 8, cfg.blended_targets.get("lpd", 0), fmt="number")
        self._write_cell(ws, 18, 9, proc.current_lpd.get("Blended", 0), fmt="number")

        # === Q1 Funnel Snapshot ===
        self._write_section_title(ws, 20, "Q1 Funnel - Deals Target vs Actual TD")
        q1_headers = ["Segment", "Q1 Target", "Actual TD", "Variance", "Attainment %"]
        self._write_headers(ws, 21, q1_headers)
        for i, seg in enumerate(["SMB", "MM", "Ent"]):
            row = 22 + i
            target = cfg.quarterly_funnel_target("Q1", seg)
            actual = proc.quarterly_funnel_actual.get("Q1", {}).get(seg, 0)
            self._write_cell(ws, row, 1, seg, bold=True)
            self._write_cell(ws, row, 2, target, fmt="number")
            self._write_cell(ws, row, 3, actual, fmt="number")
            self._write_cell(ws, row, 4, actual - target, fmt="number")
            self._write_cell(ws, row, 5, actual / target if target > 0 else 0, fmt="percent")

        # === Monthly Trend Table (Row 27) ===
        self._write_section_title(ws, 27, "Monthly Revenue Trend")
        trend_headers = ["Month", "NB Target", "NB Actual", "Exp Target", "Exp Actual", "Total"]
        self._write_headers(ws, 28, trend_headers)

        for i, m in enumerate(MONTHS):
            row = 29 + i
            nb_t = cfg.monthly_target(m, "new_biz")
            nb_a = proc.monthly_closed_won_nb[m]
            exp_t = cfg.monthly_target(m, "expansion")
            exp_a = proc.monthly_expansion_actual[m]

            self._write_cell(ws, row, 1, MONTH_NAMES[i], bold=True)
            self._write_cell(ws, row, 2, nb_t, fmt="currency")
            self._write_cell(ws, row, 3, nb_a, fmt="currency")
            self._write_cell(ws, row, 4, exp_t, fmt="currency")
            self._write_cell(ws, row, 5, exp_a, fmt="currency")
            self._write_cell(ws, row, 6, nb_a + exp_a, fmt="currency")

        # === Charts ===
        # Monthly bar chart
        cats = Reference(ws, min_col=1, min_row=29, max_row=40)
        bar_data = Reference(ws, min_col=2, min_row=28, max_row=40, max_col=3)
        chart = create_bar_chart(
            ws, "NB: Target vs Actual", bar_data, cats,
            y_title="Revenue ($)", colors=["2E75B6", "27AE60"],
        )
        ws.add_chart(chart, "H27")

        # Segment pie chart (write data for chart) — NB revenue only
        seg_start_row = 43
        self._write_cell(ws, seg_start_row, 1, "Segment", bold=True)
        self._write_cell(ws, seg_start_row, 2, "NB Revenue", bold=True)
        self._write_cell(ws, seg_start_row, 3, "Exp Revenue", bold=True)
        self._write_cell(ws, seg_start_row, 4, "Total", bold=True)
        for i, seg in enumerate(["SMB", "MM", "Ent"]):
            nb_won = proc.closed_won_nb[proc.closed_won_nb["segment"] == seg]
            exp_won = proc.closed_won_exp[proc.closed_won_exp["segment"] == seg]
            nb_rev = float(nb_won["acv"].sum())
            exp_rev = float(exp_won["acv"].sum())
            self._write_cell(ws, seg_start_row + 1 + i, 1, seg)
            self._write_cell(ws, seg_start_row + 1 + i, 2, nb_rev, fmt="currency")
            self._write_cell(ws, seg_start_row + 1 + i, 3, exp_rev, fmt="currency")
            self._write_cell(ws, seg_start_row + 1 + i, 4, nb_rev + exp_rev, fmt="currency")

        cats2 = Reference(ws, min_col=1, min_row=seg_start_row + 1, max_row=seg_start_row + 3)
        data2 = Reference(ws, min_col=2, min_row=seg_start_row, max_row=seg_start_row + 3)
        pie = create_pie_chart(
            ws, "Revenue by Segment", data2, cats2,
            colors=["2E75B6", "27AE60", "F39C12"],
        )
        ws.add_chart(pie, "H43")

    def _format(self, ws: Worksheet) -> None:
        # Set wider columns for KPI cards
        for col in range(1, 21):
            from openpyxl.utils import get_column_letter
            ws.column_dimensions[get_column_letter(col)].width = 14
        ws.sheet_properties.tabColor = "1F4E79"
