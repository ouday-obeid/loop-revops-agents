"""Excel generation wrapper — reuses generate_model.py logic for Slack bot."""

from __future__ import annotations

import traceback
from pathlib import Path

from openpyxl import Workbook

from core.config_schema import load_config
from core.loader import load_csv
from core.processor import Processor
from core.forecast_loader import load_forecast
from core.deal_matcher import match_deals
from formatting.styles import register_named_styles

from sheets.dashboard import DashboardSheet
from sheets.monthly_revenue import MonthlyRevenueSheet
from sheets.ae_scorecard import AEScorecardSheet
from sheets.pipeline_analysis import PipelineAnalysisSheet
from sheets.forecast import ForecastSheet
from sheets.segment_analysis import SegmentAnalysisSheet
from sheets.funnel_metrics import FunnelMetricsSheet
from sheets.expansion import ExpansionSheet
from sheets.sdr_performance import SDRPerformanceSheet
from sheets.monthly_unit_economics import MonthlyUnitEconomicsSheet
from sheets.raw_data import RawDataSheet
from sheets.rep_forecast import RepForecastSheet
from sheets.team_deals import TeamDealsSheet

SHEET_CLASSES = [
    DashboardSheet,
    MonthlyRevenueSheet,
    AEScorecardSheet,
    PipelineAnalysisSheet,
    ForecastSheet,
    SegmentAnalysisSheet,
    FunnelMetricsSheet,
    ExpansionSheet,
    SDRPerformanceSheet,
    MonthlyUnitEconomicsSheet,
    RawDataSheet,
    RepForecastSheet,
    TeamDealsSheet,
]


def generate_workbook(
    csv_path: str | Path | None = None,
    forecast_path: str | Path | None = None,
    config_path: str | Path | None = None,
    output_path: str | Path | None = None,
    proc: Processor | None = None,
    cfg=None,
) -> Path:
    """Generate the full revenue model Excel workbook.

    Args:
        csv_path: Path to the Salesforce CSV export.
        forecast_path: Path to the Master Forecasting Doc (optional).
        config_path: Path to the YAML config file.
        output_path: Where to save the workbook. Defaults to slack_bot/data/revenue_model.xlsx.
        proc: Pre-built Processor instance (avoids reloading CSV).
        cfg: Pre-loaded Config instance.

    Returns:
        Path to the generated Excel file.
    """
    if output_path is None:
        output_path = Path(__file__).parent / "data" / "revenue_model.xlsx"
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Use provided proc/cfg or load from files
    if proc is None or cfg is None:
        cfg = load_config(str(config_path))
        df = load_csv(str(csv_path), cfg)
        proc = Processor(df, cfg)

    # Load forecast doc if provided
    if forecast_path is not None:
        forecast_data = load_forecast(str(forecast_path), cfg)
        matched = match_deals(forecast_data, proc.df)
        proc.rep_forecast = {"data": forecast_data, "matched_deals": matched}

    # Create workbook
    wb = Workbook()
    register_named_styles(wb)

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    errors: list[str] = []
    for sheet_cls in SHEET_CLASSES:
        sheet_name = sheet_cls.sheet_name
        try:
            writer = sheet_cls(wb, cfg, proc)
            if not writer._should_generate():
                continue
            writer.generate()
        except Exception as e:
            errors.append(f"{sheet_name}: {e}")
            traceback.print_exc()
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]

    if "Dashboard" in wb.sheetnames:
        wb.active = wb.sheetnames.index("Dashboard")

    wb.save(str(output_path))

    return output_path
