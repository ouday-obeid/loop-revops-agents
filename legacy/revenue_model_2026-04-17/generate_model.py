"""CLI entry point - generates a formatted Excel revenue model from Salesforce CSV data."""

from __future__ import annotations

import argparse
import sys
import time
import traceback
from pathlib import Path

from openpyxl import Workbook

from core.config_schema import load_config
from core.loader import load_csv
from core.processor import Processor
from formatting.styles import register_named_styles

# Sheet writers in generation order (dashboard first = leftmost tab)
from sheets.dashboard import DashboardSheet
from sheets.monthly_revenue import MonthlyRevenueSheet
from sheets.ae_scorecard import AEScorecardSheet
from sheets.pipeline_analysis import PipelineAnalysisSheet
from sheets.forecast import ForecastSheet
from sheets.segment_analysis import SegmentAnalysisSheet
from sheets.funnel_metrics import FunnelMetricsSheet
from sheets.expansion import ExpansionSheet
from sheets.sdr_performance import SDRPerformanceSheet
from sheets.raw_data import RawDataSheet
from sheets.rep_forecast import RepForecastSheet

SHEET_CLASSES = [
    DashboardSheet,
    MonthlyRevenueSheet,
    AEScorecardSheet,
    PipelineAnalysisSheet,
    ForecastSheet,
    SegmentAnalysisSheet,
    FunnelMetricsSheet,
    ExpansionSheet,
    SDRPerformanceSheet,
    RawDataSheet,
    RepForecastSheet,
]


def main():
    parser = argparse.ArgumentParser(
        description="Generate a 2026 Revenue Model Excel workbook from Salesforce CSV data."
    )
    parser.add_argument("--csv", required=True, help="Path to Salesforce CSV export")
    parser.add_argument("--config", required=True, help="Path to YAML config file")
    parser.add_argument("--output", default="revenue_model_2026.xlsx",
                        help="Output Excel file path (default: revenue_model_2026.xlsx)")
    parser.add_argument("--forecast", nargs="+", default=None,
                        help="Path(s) to Master Forecasting Doc Excel file(s) (optional, accepts multiple)")
    parser.add_argument("--verbose", action="store_true", help="Print detailed progress")
    args = parser.parse_args()

    start_time = time.time()
    warnings_count = 0
    skipped_sheets = []

    # --- Load config ---
    print(f"Loading config: {args.config}")
    cfg = load_config(args.config)
    print(f"  Fiscal year: {cfg.fiscal_year}")
    print(f"  AE roster: {len(cfg.ae_roster)} AEs")

    # --- Load CSV ---
    print(f"\nLoading CSV: {args.csv}")
    df = load_csv(args.csv, cfg)
    print(f"  Loaded {len(df)} opportunities")
    print(f"  Date range: {df['close_date'].min()} to {df['close_date'].max()}")
    print(f"  Stages: {df['stage'].nunique()} unique")

    # --- Initialize processor ---
    proc = Processor(df, cfg)

    # --- Load forecast doc (optional) ---
    if args.forecast:
        from core.forecast_loader import load_forecast, load_forecasts
        from core.deal_matcher import match_deals

        if len(args.forecast) == 1:
            forecast_data = load_forecast(args.forecast[0], cfg, verbose=args.verbose)
        else:
            forecast_data = load_forecasts(args.forecast, cfg, verbose=args.verbose)
        matched = match_deals(forecast_data, df, verbose=args.verbose)
        proc.rep_forecast = {"data": forecast_data, "matched_deals": matched}
        print(f"  Rep forecast loaded: {len(forecast_data.reps)} reps")

    if args.verbose:
        print(f"\n--- Quick Stats ---")
        print(f"  Closed Won: {proc.ytd_closed_won_count} deals, ${proc.ytd_closed_won_acv:,.0f}")
        print(f"  NB Revenue: ${proc.ytd_closed_won_nb_acv:,.0f}")
        print(f"  Expansion Revenue: ${proc.ytd_closed_won_exp_acv:,.0f}")
        print(f"  Open Pipeline: ${proc.total_pipeline_acv:,.0f}")
        print(f"  Weighted Pipeline: ${proc.total_weighted_pipeline:,.0f}")
        print(f"  Win Rate: {proc.overall_win_rate:.1%}")

    # --- Create workbook ---
    print(f"\nGenerating workbook...")
    wb = Workbook()
    register_named_styles(wb)

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # --- Generate each sheet ---
    for sheet_cls in SHEET_CLASSES:
        sheet_name = sheet_cls.sheet_name
        try:
            writer = sheet_cls(wb, cfg, proc)
            if not writer._should_generate():
                if args.verbose:
                    print(f"  Skipping: {sheet_name} (no data)")
                continue
            if args.verbose:
                print(f"  Writing: {sheet_name}...", end=" ", flush=True)
            writer.generate()
            if args.verbose:
                print("OK")
        except Exception as e:
            skipped_sheets.append(sheet_name)
            warnings_count += 1
            print(f"  ERROR writing '{sheet_name}': {e}", file=sys.stderr)
            if args.verbose:
                traceback.print_exc()
            # Remove the broken sheet if it was created
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]

    # --- Set Dashboard as active ---
    if "Dashboard" in wb.sheetnames:
        wb.active = wb.sheetnames.index("Dashboard")

    # --- Save ---
    output_path = Path(args.output)
    print(f"\nSaving to: {output_path.absolute()}")
    wb.save(str(output_path))

    elapsed = time.time() - start_time

    # --- Summary ---
    print(f"\n{'='*50}")
    print(f"  COMPLETE in {elapsed:.1f}s")
    print(f"  Sheets generated: {len(wb.sheetnames)}")
    if skipped_sheets:
        print(f"  Skipped sheets: {', '.join(skipped_sheets)}")
    if warnings_count:
        print(f"  Warnings: {warnings_count}")
    print(f"  Output: {output_path.absolute()}")
    print(f"{'='*50}")


if __name__ == "__main__":
    main()
