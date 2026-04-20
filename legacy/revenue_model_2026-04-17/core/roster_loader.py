"""Load Sales Team Roster from Excel and update Config rosters dynamically."""

from __future__ import annotations

import sys
from pathlib import Path

import openpyxl

from core.config_schema import Config


# Map the "Level" column in the roster to segment codes
_LEVEL_TO_SEGMENT = {
    "mid market": "MM",
    "mm": "MM",
    "smb": "SMB",
    "ent": "Ent",
    "enterprise": "Ent",
}

# Map Role column to internal role values
_ROLE_MAP = {
    "ae": "AE",
    "sdr": "SDR",
    "sdr team lead": "SDR Team Lead",
    "manager": "Manager",
}

# Default quotas by segment (used when roster doesn't specify)
_DEFAULT_QUOTA = {
    "SMB": 1200000,
    "MM": 1000000,
    "Ent": 1400000,
}

# Manager full name → group alias used in digest/config
# Maps the Manager column value in the roster to the short group name
_MANAGER_ALIAS = {
    "Arthur Fisher": "Hutch",
    "Nathan Meyer": "Nate",
    "Henry Gimelfarb": "IC",
    "Charles Kagahastian": "Charles",
}

# Roster name → Salesforce name mapping for known mismatches
_NAME_ALIAS = {
    "Clayton Arvizu": "Clay Arvizu",
}


def load_roster(path: str | Path, cfg: Config) -> dict:
    """Parse the Sales Team Roster Excel and return roster data.

    Returns a dict with keys:
        ae_roster: list of AE dicts
        sdr_roster: list of SDR dicts
        manager_groups: dict mapping manager alias to list of AE names
    """
    path = Path(path)
    if not path.exists():
        print(f"WARNING: Roster file not found: {path}", file=sys.stderr)
        return {}

    wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)

    # Find the AE Team sheet
    ae_ws = None
    for name in wb.sheetnames:
        if name.lower().startswith("ae team"):
            ae_ws = wb[name]
            break
    if ae_ws is None:
        print("WARNING: No 'AE Team' sheet found in roster file", file=sys.stderr)
        wb.close()
        return {}

    # Parse header row (row 2) to find column indices
    rows = list(ae_ws.iter_rows(min_row=1, values_only=True))
    wb.close()

    if len(rows) < 3:
        print("WARNING: Roster sheet has too few rows", file=sys.stderr)
        return {}

    # Row 2 (index 1) has the actual column headers
    headers = [str(h).strip().lower() if h else "" for h in rows[1]]

    def _col(name):
        for i, h in enumerate(headers):
            if h == name:
                return i
        return -1

    col_name = _col("name")
    col_role = _col("role")
    col_level = _col("level")
    col_manager = _col("manager")
    col_status = _col("status")

    if col_name < 0:
        print("WARNING: Could not find 'Name' column in roster", file=sys.stderr)
        return {}

    ae_roster = []
    sdr_roster = []
    manager_groups: dict[str, list[str]] = {}

    # Build existing quota lookup from current config for preservation
    existing_quotas = {}
    for ae in cfg.ae_roster:
        existing_quotas[ae["name"]] = ae.get("quota", 0)

    for row in rows[2:]:  # Skip header rows
        if not row or not row[col_name]:
            continue

        raw_name = str(row[col_name]).strip()
        if not raw_name:
            continue

        # Apply name alias for SF matching
        name = _NAME_ALIAS.get(raw_name, raw_name)

        role_raw = str(row[col_role]).strip().lower() if col_role >= 0 and row[col_role] else "ae"
        role = _ROLE_MAP.get(role_raw, role_raw.title())

        level_raw = str(row[col_level]).strip().lower() if col_level >= 0 and row[col_level] else ""
        segment = _LEVEL_TO_SEGMENT.get(level_raw, "MM")

        manager_full = str(row[col_manager]).strip() if col_manager >= 0 and row[col_manager] else ""

        status = str(row[col_status]).strip().lower() if col_status >= 0 and row[col_status] else "ramping"

        if role in ("AE",):
            # Check quota by both raw_name and aliased name
            quota = existing_quotas.get(name,
                    existing_quotas.get(raw_name,
                    _DEFAULT_QUOTA.get(segment, 1000000)))
            ae_entry = {
                "name": name,
                "quota": quota,
                "segment": segment,
                "status": status,
            }
            ae_roster.append(ae_entry)

            # Build manager groups using alias
            if manager_full:
                mgr_alias = _MANAGER_ALIAS.get(manager_full, manager_full.split()[0])
                if mgr_alias not in manager_groups:
                    manager_groups[mgr_alias] = []
                manager_groups[mgr_alias].append(name)

        elif role in ("Manager",):
            ae_entry = {
                "name": name,
                "quota": existing_quotas.get(name, 0),
                "segment": segment or "Ent",
                "status": status,
                "role": "Manager",
            }
            ae_roster.append(ae_entry)

        elif role in ("SDR", "SDR Team Lead"):
            sdr_entry = {
                "name": name,
                "segment": segment,
                "status": status,
                "role": role,
            }
            sdr_roster.append(sdr_entry)

    return {
        "ae_roster": ae_roster,
        "sdr_roster": sdr_roster,
        "manager_groups": manager_groups,
    }


def apply_roster(cfg: Config, roster_data: dict) -> None:
    """Apply roster data to a Config object, overriding static YAML values."""
    if not roster_data:
        return

    if "ae_roster" in roster_data and roster_data["ae_roster"]:
        cfg.ae_roster = roster_data["ae_roster"]
    if "sdr_roster" in roster_data and roster_data["sdr_roster"]:
        cfg.sdr_roster = roster_data["sdr_roster"]
    if "manager_groups" in roster_data and roster_data["manager_groups"]:
        cfg.manager_groups = roster_data["manager_groups"]
