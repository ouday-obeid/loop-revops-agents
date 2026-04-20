"""Parse the Master Forecasting Doc Excel file for rep-driven forecast data."""

from __future__ import annotations

import sys
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl

from core.config_schema import Config


@dataclass
class RepDeal:
    """A single deal from a rep's forecast tab."""
    name: str
    acv: float
    tier: str  # "Commit", "HC", or "Longshot"


@dataclass
class RepForecast:
    """Forecast data for a single rep."""
    short_name: str
    full_name: str = ""
    current_arr: float = 0.0
    pipeline: float = 0.0
    in_proposal: float = 0.0
    commit_total: float = 0.0
    hc_total: float = 0.0
    longshot_total: float = 0.0
    quota: float = 0.0
    deals: list[RepDeal] = field(default_factory=list)
    has_detail_tab: bool = False


@dataclass
class RepForecastData:
    """All rep forecast data parsed from the Master Forecasting Doc."""
    reps: dict[str, RepForecast] = field(default_factory=dict)
    manager_sections: dict[str, list[str]] = field(default_factory=dict)


# Nickname → first name mapping for forecast doc resolution
_NICKNAMES = {
    "nicky b": "Nick",
    "dan": "Daniel",
}


def _resolve_name(short_name: str, ae_roster: list[dict]) -> str:
    """Resolve short name ('Alexis') to full config name ('Alexis Marrero')
    by first-name prefix match against the roster."""
    short_lower = short_name.strip().lower()
    # Check nickname mapping first
    canonical = _NICKNAMES.get(short_lower, short_name).strip().lower()
    for ae in ae_roster:
        full = ae["name"]
        first = full.split()[0].lower()
        if first == canonical or first == short_lower:
            return full
    return short_name


def _parse_currency(val) -> float:
    """Parse a cell value to float, handling currency strings and None."""
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).replace("$", "").replace(",", "").strip()
    if not s or s == "-":
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def _parse_master_sheet(ws, cfg: Config) -> RepForecastData:
    """Parse the Master sheet for per-rep summary rows.

    Layout (based on plan):
      Col B = Current ARR
      Col D = Pipeline
      Col E = In Proposal
      Col G = Commit total
      Col H = HC total
      Col I = Longshot total
      Col M = Quota

    Manager groupings detected by rows ending with ':' (e.g. "Nate:")
    """
    data = RepForecastData()
    current_manager = None

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        cell_a = row[0].value  # Column A
        if cell_a is None:
            continue

        cell_a_str = str(cell_a).strip()

        # Detect manager grouping rows (e.g. "Nate:", "Charles:")
        if cell_a_str.endswith(":"):
            current_manager = cell_a_str.rstrip(":")
            if current_manager not in data.manager_sections:
                data.manager_sections[current_manager] = []
            continue

        # Skip header-like or total rows
        if cell_a_str.lower() in ("", "name", "total", "grand total", "rep"):
            continue

        # Try to resolve as an AE name
        full_name = _resolve_name(cell_a_str, cfg.ae_roster)

        # Only include if the resolved name is in the AE roster
        if full_name not in cfg.ae_only_names:
            continue

        # Parse columns (0-indexed in the row tuple)
        rf = RepForecast(
            short_name=cell_a_str,
            full_name=full_name,
            current_arr=_parse_currency(row[1].value if len(row) > 1 else None),   # B
            pipeline=_parse_currency(row[3].value if len(row) > 3 else None),       # D
            in_proposal=_parse_currency(row[4].value if len(row) > 4 else None),    # E
            commit_total=_parse_currency(row[6].value if len(row) > 6 else None),   # G
            hc_total=_parse_currency(row[7].value if len(row) > 7 else None),       # H
            longshot_total=_parse_currency(row[8].value if len(row) > 8 else None), # I
            quota=_parse_currency(row[12].value if len(row) > 12 else None),        # M
        )

        data.reps[full_name] = rf

        if current_manager is not None:
            data.manager_sections[current_manager].append(full_name)

    return data


def _parse_rep_tab(ws, rep_name: str) -> list[RepDeal]:
    """Parse a per-rep forecast tab (e.g. 'Alexis Forecast').

    Expected layout: 3 sections (Commit / HC / Longshot), each with
    deal name + ACV columns. Sections are detected by header rows
    containing 'commit', 'minimum', 'high commit' / 'hc', or 'longshot'.

    Handles two column layouts:
      - Layout 1: col A = deal name, col B = ACV
      - Layout 2: col B = deal name, col C = ACV (col A blank or label)
    """
    deals: list[RepDeal] = []
    current_tier = None

    # Detect layout: if col A is mostly blank/labels and col B has names, use layout 2
    use_bc = False
    max_row = ws.max_row or 500
    for row in ws.iter_rows(min_row=1, max_row=min(5, max_row), values_only=False):
        if len(row) > 1 and row[1].value and str(row[1].value).strip().lower() in ("name", "name of account (sfdc link)"):
            use_bc = True
            break

    for row in ws.iter_rows(min_row=1, max_row=max_row, values_only=False):
        # Check both col A and col B for tier headers
        tier_text = None
        for idx in (0, 1):
            if idx < len(row) and row[idx].value:
                tier_text = str(row[idx].value).strip().lower()
                break
        if tier_text is None:
            continue

        # Detect tier headers — "Minimum" maps to "Commit"
        if tier_text in ("commit", "minimum") or (tier_text.startswith("commit") and "high" not in tier_text):
            current_tier = "Commit"
            continue
        elif "high commit" in tier_text or "high commitment" in tier_text or tier_text == "hc":
            current_tier = "HC"
            continue
        elif "longshot" in tier_text:
            current_tier = "Longshot"
            continue

        # Skip header/label rows within a tier
        if current_tier is None:
            continue

        # Extract deal name and ACV based on layout
        if use_bc:
            name_val = row[1].value if len(row) > 1 else None
            acv_val = row[2].value if len(row) > 2 else None
        else:
            name_val = row[0].value
            acv_val = row[1].value if len(row) > 1 else None

        if name_val is None:
            continue
        name_str = str(name_val).strip().lower()
        if name_str in ("deal", "account", "name", "deal name", "total", "",
                         "name of account (sfdc link)"):
            continue

        deal_name = str(name_val).strip()
        acv = _parse_currency(acv_val)

        if deal_name and acv > 0:
            deals.append(RepDeal(name=deal_name, acv=acv, tier=current_tier))

    return deals


def _load_single_forecast(path: Path, cfg: Config, verbose: bool = False) -> RepForecastData:
    """Load and parse a single forecast doc file."""
    if verbose:
        print(f"\nLoading forecast doc: {path}")

    wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)

    # Find the Master sheet (first sheet, or named "Master" / "Summary")
    master_ws = None
    for name in wb.sheetnames:
        if name.lower() in ("master", "summary", "team",
                             "master forecasting sheet"):
            master_ws = wb[name]
            break
    if master_ws is None:
        master_ws = wb[wb.sheetnames[0]]

    if verbose:
        print(f"  Parsing master sheet: '{master_ws.title}'")

    data = _parse_master_sheet(master_ws, cfg)

    if verbose:
        print(f"  Found {len(data.reps)} reps on master sheet")

    # Look for per-rep detail tabs (e.g. "Alexis Forecast", "James Forecast",
    # "Simon Deal Forecast")
    for sheet_name in wb.sheetnames:
        if "forecast" not in sheet_name.lower():
            continue
        if sheet_name == master_ws.title:
            continue

        # Extract the rep first name from the tab name
        # e.g. "Alexis Forecast" -> "Alexis", "Simon Deal Forecast" -> "Simon"
        parts = sheet_name.split()
        if len(parts) < 2:
            continue
        rep_short = parts[0]
        rep_full = _resolve_name(rep_short, cfg.ae_roster)

        if rep_full in data.reps:
            ws = wb[sheet_name]
            deals = _parse_rep_tab(ws, rep_full)
            data.reps[rep_full].deals = deals
            data.reps[rep_full].has_detail_tab = True
            if verbose:
                print(f"  Parsed {len(deals)} deals from '{sheet_name}' for {rep_full}")
        elif rep_full in cfg.ae_only_names:
            # Rep has a tab but wasn't on master sheet — create entry from tab
            rf = RepForecast(short_name=rep_short, full_name=rep_full, has_detail_tab=True)
            ws = wb[sheet_name]
            rf.deals = _parse_rep_tab(ws, rep_full)
            # Compute tier totals from deals
            rf.commit_total = sum(d.acv for d in rf.deals if d.tier == "Commit")
            rf.hc_total = sum(d.acv for d in rf.deals if d.tier == "HC")
            rf.longshot_total = sum(d.acv for d in rf.deals if d.tier == "Longshot")
            data.reps[rep_full] = rf
            if verbose:
                print(f"  Parsed {len(rf.deals)} deals from '{sheet_name}' for {rep_full} (tab only)")

    wb.close()

    if verbose:
        total_deals = sum(len(r.deals) for r in data.reps.values())
        detail_reps = sum(1 for r in data.reps.values() if r.has_detail_tab)
        print(f"  Total: {len(data.reps)} reps, {detail_reps} with deal detail, {total_deals} deals")

    return data


def _merge_forecast_data(all_data: list[RepForecastData]) -> RepForecastData:
    """Merge multiple RepForecastData objects into one.

    For reps that appear in multiple files, the version with a detail tab
    (individual deals) takes priority. If both have detail tabs, the one
    with more deals wins. For master-sheet-only data, later files overwrite
    earlier ones.
    """
    merged = RepForecastData()

    for data in all_data:
        for mgr, reps in data.manager_sections.items():
            if mgr not in merged.manager_sections:
                merged.manager_sections[mgr] = []
            for rep in reps:
                if rep not in merged.manager_sections[mgr]:
                    merged.manager_sections[mgr].append(rep)

        for name, rf in data.reps.items():
            existing = merged.reps.get(name)
            if existing is None:
                merged.reps[name] = rf
            elif rf.has_detail_tab and not existing.has_detail_tab:
                merged.reps[name] = rf
            elif rf.has_detail_tab and existing.has_detail_tab:
                # Keep the one with more deals
                if len(rf.deals) >= len(existing.deals):
                    merged.reps[name] = rf
            # else: existing has detail tab and new doesn't — keep existing

    return merged


def load_forecast(path: str | Path, cfg: Config, verbose: bool = False) -> RepForecastData:
    """Load and parse a single Master Forecasting Doc.

    Args:
        path: Path to the Excel file.
        cfg: Config object for name resolution.
        verbose: Print progress info.

    Returns:
        RepForecastData with per-rep forecast data.
    """
    path = Path(path)
    if not path.exists():
        print(f"ERROR: Forecast file not found: {path}", file=sys.stderr)
        sys.exit(1)

    return _load_single_forecast(path, cfg, verbose=verbose)


def load_forecasts(paths: list[str | Path], cfg: Config, verbose: bool = False) -> RepForecastData:
    """Load and merge multiple forecast doc files.

    Args:
        paths: List of paths to Excel forecast files.
        cfg: Config object for name resolution.
        verbose: Print progress info.

    Returns:
        Merged RepForecastData with per-rep forecast data from all files.
    """
    all_data = []
    for p in paths:
        p = Path(p)
        if not p.exists():
            print(f"WARNING: Forecast file not found, skipping: {p}", file=sys.stderr)
            continue
        all_data.append(_load_single_forecast(p, cfg, verbose=verbose))

    if not all_data:
        print("ERROR: No valid forecast files found.", file=sys.stderr)
        sys.exit(1)

    merged = _merge_forecast_data(all_data)

    if verbose:
        total_deals = sum(len(r.deals) for r in merged.reps.values())
        detail_reps = sum(1 for r in merged.reps.values() if r.has_detail_tab)
        print(f"\n  Merged forecast: {len(merged.reps)} reps, "
              f"{detail_reps} with deal detail, {total_deals} deals")

    return merged
