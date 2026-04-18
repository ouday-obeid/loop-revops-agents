"""Utility functions: auto-width, freeze panes, merge helpers."""

from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from formatting.styles import (
    ALIGN_CENTER, ALIGN_RIGHT, FILL_HEADER, FONT_HEADER,
    FONT_TITLE, THIN_BORDER, FMT_CURRENCY, FMT_PERCENT, FMT_NUMBER,
)


def auto_width(ws: Worksheet, min_width: int = 8, max_width: int = 50):
    """Auto-size columns based on content."""
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = min_width
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx,
                                min_row=1, max_row=min(ws.max_row, 200)):
            cell = row[0]
            if cell.value is not None:
                text = str(cell.value)
                max_len = max(max_len, len(text) + 2)
        ws.column_dimensions[col_letter].width = min(max_len, max_width)


def freeze_panes(ws: Worksheet, row: int = 2, col: int = 1):
    """Freeze panes at the given row/col (1-indexed)."""
    col_letter = get_column_letter(col)
    ws.freeze_panes = f"{col_letter}{row}"


def write_header_row(ws: Worksheet, row: int, headers: list[str], start_col: int = 1):
    """Write a styled header row."""
    for i, header in enumerate(headers):
        cell = ws.cell(row=row, column=start_col + i, value=header)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER


def write_title(ws: Worksheet, row: int, title: str, col: int = 1):
    """Write a section title."""
    cell = ws.cell(row=row, column=col, value=title)
    cell.font = FONT_TITLE


def write_data_rows(ws: Worksheet, start_row: int, data: list[list],
                    formats: list[str] | None = None, start_col: int = 1):
    """Write rows of data with optional per-column formatting.

    formats: list of 'currency', 'percent', 'number', or None per column.
    """
    fmt_map = {
        "currency": FMT_CURRENCY,
        "percent": FMT_PERCENT,
        "number": FMT_NUMBER,
    }
    align_map = {
        "currency": ALIGN_RIGHT,
        "percent": ALIGN_RIGHT,
        "number": ALIGN_RIGHT,
    }
    for r_idx, row_data in enumerate(data):
        for c_idx, value in enumerate(row_data):
            cell = ws.cell(row=start_row + r_idx, column=start_col + c_idx, value=value)
            if formats and c_idx < len(formats) and formats[c_idx]:
                fmt_key = formats[c_idx]
                if fmt_key in fmt_map:
                    cell.number_format = fmt_map[fmt_key]
                if fmt_key in align_map:
                    cell.alignment = align_map[fmt_key]


def merge_and_center(ws: Worksheet, start_row: int, start_col: int,
                     end_row: int, end_col: int, value: str = ""):
    """Merge cells and center the value."""
    ws.merge_cells(
        start_row=start_row, start_column=start_col,
        end_row=end_row, end_column=end_col,
    )
    cell = ws.cell(row=start_row, column=start_col, value=value)
    cell.alignment = ALIGN_CENTER
    return cell
