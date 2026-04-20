"""Conditional formatting rule builders."""

from openpyxl.formatting.rule import CellIsRule, DataBarRule
from openpyxl.styles import Font, PatternFill

# Attainment fills
_GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_GREEN_FONT = Font(color="006100")
_YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
_YELLOW_FONT = Font(color="9C5700")
_RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_RED_FONT = Font(color="9C0006")


def add_attainment_formatting(ws, cell_range: str):
    """3-tier attainment: green >=100%, yellow >=80%, red <80%."""
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(operator="greaterThanOrEqual", formula=["1.0"],
                   fill=_GREEN_FILL, font=_GREEN_FONT),
    )
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(operator="between", formula=["0.8", "0.9999"],
                   fill=_YELLOW_FILL, font=_YELLOW_FONT),
    )
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(operator="lessThan", formula=["0.8"],
                   fill=_RED_FILL, font=_RED_FONT),
    )


def add_variance_formatting(ws, cell_range: str):
    """Red for negative variance, green for positive."""
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(operator="lessThan", formula=["0"],
                   fill=_RED_FILL, font=_RED_FONT),
    )
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(operator="greaterThanOrEqual", formula=["0"],
                   fill=_GREEN_FILL, font=_GREEN_FONT),
    )


def add_data_bars(ws, cell_range: str, color: str = "2E75B6"):
    """Add data bars to a range."""
    rule = DataBarRule(
        start_type="min", end_type="max",
        color=color, showValue=True,
    )
    ws.conditional_formatting.add(cell_range, rule)
