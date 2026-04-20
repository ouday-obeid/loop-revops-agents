"""Color palette, named styles, and font definitions."""

from openpyxl.styles import (
    Alignment, Border, Font, NamedStyle, PatternFill, Side,
)

# --- Color Palette ---
NAVY = "1F4E79"
DARK_BLUE = "2E75B6"
LIGHT_BLUE = "D6E4F0"
WHITE = "FFFFFF"
LIGHT_GRAY = "F2F2F2"
GREEN = "27AE60"
YELLOW = "F39C12"
RED = "E74C3C"
DARK_GRAY = "4A4A4A"

# --- Fonts ---
FONT_HEADER = Font(name="Calibri", size=11, bold=True, color=WHITE)
FONT_SUBHEADER = Font(name="Calibri", size=11, bold=True, color=NAVY)
FONT_TITLE = Font(name="Calibri", size=14, bold=True, color=NAVY)
FONT_BODY = Font(name="Calibri", size=10, color=DARK_GRAY)
FONT_BODY_BOLD = Font(name="Calibri", size=10, bold=True, color=DARK_GRAY)
FONT_KPI_VALUE = Font(name="Calibri", size=18, bold=True, color=NAVY)
FONT_KPI_LABEL = Font(name="Calibri", size=9, color=DARK_GRAY)

# --- Fills ---
FILL_HEADER = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
FILL_SUBHEADER = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
FILL_ALT_ROW = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
FILL_GREEN = PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")
FILL_YELLOW = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type="solid")
FILL_RED = PatternFill(start_color=RED, end_color=RED, fill_type="solid")
FILL_WHITE = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
FILL_KPI_BG = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")

# --- Borders ---
THIN_BORDER = Border(
    left=Side(style="thin", color=DARK_BLUE),
    right=Side(style="thin", color=DARK_BLUE),
    top=Side(style="thin", color=DARK_BLUE),
    bottom=Side(style="thin", color=DARK_BLUE),
)
BOTTOM_BORDER = Border(bottom=Side(style="medium", color=NAVY))

# --- Alignment ---
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

# --- Number Formats ---
FMT_CURRENCY = '$#,##0'
FMT_CURRENCY_K = '$#,##0,"K"'
FMT_PERCENT = '0.0%'
FMT_NUMBER = '#,##0'
FMT_DECIMAL = '#,##0.0'


def register_named_styles(wb):
    """Register reusable named styles on the workbook."""
    styles = {
        "header": _make_style("header", FONT_HEADER, FILL_HEADER, ALIGN_CENTER, THIN_BORDER),
        "subheader": _make_style("subheader", FONT_SUBHEADER, FILL_SUBHEADER, ALIGN_CENTER, THIN_BORDER),
        "body": _make_style("body", FONT_BODY, None, ALIGN_LEFT, None),
        "currency": _make_style("currency", FONT_BODY, None, ALIGN_RIGHT, None, FMT_CURRENCY),
        "percent": _make_style("percent", FONT_BODY, None, ALIGN_RIGHT, None, FMT_PERCENT),
        "number": _make_style("number", FONT_BODY, None, ALIGN_RIGHT, None, FMT_NUMBER),
    }
    for name, style in styles.items():
        try:
            wb.add_named_style(style)
        except ValueError:
            pass  # Already registered


def _make_style(name, font, fill, alignment, border, number_format=None):
    ns = NamedStyle(name=name)
    ns.font = font
    if fill:
        ns.fill = fill
    ns.alignment = alignment
    if border:
        ns.border = border
    if number_format:
        ns.number_format = number_format
    return ns
